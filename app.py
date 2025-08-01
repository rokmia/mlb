import requests
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def get_all_team_ids():
    url = "https://statsapi.mlb.com/api/v1/teams?sportId=1"
    teams = requests.get(url).json()["teams"]
    return [(team["id"], team["name"]) for team in teams]

def get_team_roster(team_id):
    url = f"https://statsapi.mlb.com/api/v1/teams/{team_id}/roster"
    roster = requests.get(url).json()
    return [player["person"]["id"] for player in roster["roster"]]

def get_player_info_and_stats(player_id, team_name):
    url = f"https://statsapi.mlb.com/api/v1/people/{player_id}?hydrate=stats(group=[hitting],type=[season,career])"
    response = requests.get(url).json()
    person = response["people"][0]

    name = person["fullName"]
    jersey = person.get("primaryNumber", "N/A")
    birthdate = person.get("birthDate", "N/A")

    stats = {
        "Season HR": None,
        "Career HR": None,
        "Season Games": None,
        "Career Games": None,
        "Season Hits": None,
        "Career Hits": None,
        "Season Runs": None,
        "Career Runs": None,
        "Season RBI": None,
        "Career RBI": None,
        "Season TB": None,
        "Career TB": None,
    }

    for stat_block in person.get("stats", []):
        stat_type = stat_block["type"]["displayName"]
        splits = stat_block.get("splits", [])
        if not splits:
            continue

        stat_data = splits[0]["stat"]
        prefix = "Season" if stat_type == "season" else "Career"
        stats[f"{prefix} HR"] = int(stat_data.get("homeRuns", 0))
        stats[f"{prefix} Games"] = int(stat_data.get("gamesPlayed", 0))
        stats[f"{prefix} Hits"] = int(stat_data.get("hits", 0))
        stats[f"{prefix} Runs"] = int(stat_data.get("runs", 0))
        stats[f"{prefix} RBI"] = int(stat_data.get("rbi", 0))
        stats[f"{prefix} TB"] = int(stat_data.get("totalBases", 0))

    return {
        "Name": name,
        "Team": team_name,
        "Jersey Number": jersey,
        "Birth Date": birthdate,
        **stats
    }

def get_players_near_13_all_teams():
    all_results = []
    for team_id, team_name in get_all_team_ids():
        try:
            player_ids = get_team_roster(team_id)
        except:
            continue

        for pid in player_ids:
            try:
                player = get_player_info_and_stats(pid, team_name)
            except:
                continue

            trigger_notes = []
            for stat_key in [
                "Season HR", "Career HR",
                "Season Games", "Career Games",
                "Season Hits", "Career Hits",
                "Season Runs", "Career Runs",
                "Season RBI", "Career RBI",
                "Season TB", "Career TB"
            ]:
                value = player.get(stat_key)
                if value is not None and value % 13 == 12:
                    next_milestone = value + 1
                    trigger_notes.append(f"{stat_key} = {value} (1 away from {next_milestone})")

            if trigger_notes:
                player["Milestone Trigger"] = "; ".join(trigger_notes)
                all_results.append(player)

            time.sleep(0.3)

    return all_results

def export_to_excel(data, filename):
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Color milestone column yellow
    yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column_letter == 'P':  # Assuming 'Milestone Trigger' is column P (adjust if needed)
                if cell.value:
                    cell.fill = yellow_fill

    # Autofit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

# Run script
players = get_players_near_13_all_teams()
if players:
    export_to_excel(players, "milestone_watch_13.xlsx")
    print("✅ Exported to milestone_watch_13.xlsx with highlights.")
else:
    print("⚠️ No players found near any 13 milestone.")
